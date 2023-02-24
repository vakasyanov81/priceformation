# -*- coding: utf-8 -*-
"""
xls read logic
"""

__author__ = "Kasyanov V.A."

import os
from typing import List

import xlrd
from openpyxl import load_workbook

import core
from cfg import init_cfg

_MAX_COLUMNS = 50
_MAX_ROWS = 10000
__SKIPPED_EMPTY_ROW__ = 10

cfg = init_cfg()


class IXlsReader:
    """interface xls reader"""

    # pylint: disable=R0903

    def parse(self, sheet_indexes: list = None):
        """do parse"""
        raise NotImplementedError


class FakeXlsReader:
    """fake xls reader"""

    parse_result = None

    @classmethod
    def get_instance(cls, file_path, params):
        """get instance FakeXlsReader"""
        return FakeXlsReader(file_path, params)

    def __init__(self, file_path, params):
        """init"""
        self.file_path = file_path
        self.params = params
        self.sheet_indexes = None

    def parse(self, sheet_indexes: list = None):
        """do parse"""
        # pylint: disable=E1102

        self.sheet_indexes = sheet_indexes
        return self.parse_result() if callable(self.parse_result) else self.parse_result


class ParamsHelper:
    """params data container"""

    params = {}

    @property
    def start_row(self) -> int:
        """start row index"""
        return self.params.get("start_row")

    @property
    def cur_row(self) -> int:
        """current process row"""
        return self.params.get("cur_row") or self.start_row

    @cur_row.setter
    def cur_row(self, cur_row: int):
        """current process row setter"""
        self.params["cur_row"] = cur_row

    @property
    def columns(self) -> dict:
        """mapping columns. {0: "title", 1: "price"...}"""
        return self.params.get("columns")

    @property
    def max_columns(self) -> int:
        """max columns to be processed"""
        return self.params.get("max_columns") or _MAX_COLUMNS

    @property
    def max_rows(self) -> int:
        """max rows to be processed"""
        return self.params.get("max_rows") or _MAX_ROWS


class XlsReader(IXlsReader, ParamsHelper):
    """xls reader"""

    @classmethod
    def get_instance(cls, file_path, params):
        """get instance XlsReader / XlsxReader"""
        if ".xlsx" in file_path:
            return XlsxReader(file_path, params)
        return XlsReader(file_path, params)

    def __init__(self, file_path, params):
        """init"""
        self.cur_row_values = []
        self._skipped_empty_rows = 0
        self.book = None
        self.params: dict = params
        self.open_book(file_path)

    def open_book(self, file_path):
        """open book"""
        file_path = f"{cfg.main.project_root}{os.sep}{file_path}"
        self.book = self._open_book(file_path)

    @classmethod
    def _open_book(cls, file_path):
        """open book"""
        return xlrd.open_workbook(file_path)

    def skipped_rows(self):
        """get skipped rows count value"""
        return self._skipped_empty_rows

    def get_sheet_names(self) -> str:
        """get sheet names"""
        return self.book.sheet_names()

    def get_sheet_by_name(self, s_name):
        """get sheet by name"""
        return self.book.sheet_by_name(s_name)

    def sheets(self) -> list:
        """get sheet list"""

        if not self.get_sheet_names():
            core.make_raise("В прайсе отсутствуют вкладки!")

        return [self.get_sheet_by_name(s_name) for s_name in self.get_sheet_names()]

    def next_row_values(self, sheet):
        """process for next xls row"""

        def _(val):
            return val.strip() if isinstance(val, str) else val

        if self.is_end_row():
            return None

        end_col = (
            self.sheet_cols(sheet)
            if self.sheet_cols(sheet) <= self.max_columns
            else self.max_columns
        )
        cur_row = self.cur_row
        self.cur_row += 1

        if self.cur_row > self.max_rows:
            raise MaxRowsReached(self.max_rows)

        try:
            self.cur_row_values = [
                _(cell) for cell in self.row_values(sheet, cur_row, end_col)
            ]
        except IndexError:
            self.cur_row_values = [None]

        if self.is_empty_row():
            self._skipped_empty_rows += 1
            return not self.is_end_row()

        return self.cur_row_values

    @classmethod
    def row_values(cls, sheet: "xlrd.sheet.Sheet", cur_row, end_col):
        """get row values"""
        return sheet.row_values(cur_row, end_colx=end_col)

    @classmethod
    def sheet_cols(cls, sheet):
        """number of columns on sheet"""
        return sheet.ncols

    def is_empty_row(self):
        """row is empty?"""
        for value in self.cur_row_values:
            if value:
                return False
        return True

    def is_end_row(self):
        """is end row?"""
        return self.is_empty_row() and self._skipped_empty_rows >= __SKIPPED_EMPTY_ROW__

    def parse(self, sheet_indexes: list = None):
        """parse given sheets or all if not specified"""
        rows = []

        all_num_sheets = len(self.sheets())

        sheet_indexes = sheet_indexes or list(range(0, all_num_sheets))

        for sheet_index in sheet_indexes:
            self._skipped_empty_rows = 0
            self.cur_row = self.start_row
            rows += self.parse_sheet(self.sheets()[sheet_index])

        return rows

    def parse_sheet(self, sheet) -> List[dict]:
        """parse sheet"""
        rows = []
        while self.next_row_values(sheet):
            if self.is_empty_row():
                continue
            rows.append(self.to_dict(self.cur_row_values))

        return rows

    def to_dict(self, row):
        """row to dict"""

        result = {}
        col_numbers = list(self.columns.keys())
        for col_number in col_numbers:
            val = row[col_number]
            col_name = self.columns.get(col_number)
            result[col_name] = val

        return result


class XlsxReader(XlsReader):
    """xlsx reader"""

    @classmethod
    def _open_book(cls, file_path):
        """open book"""
        return load_workbook(file_path)

    def get_sheet_names(self) -> str:
        """get sheet names"""
        return self.book.sheetnames

    def get_sheet_by_name(self, s_name):
        """get sheet by name"""
        return self.book[s_name]

    @classmethod
    def sheet_cols(cls, sheet):
        return sheet.max_column

    @classmethod
    def row_values(cls, sheet, cur_row, end_col):
        values = []
        for i in range(end_col):
            values.append(sheet.cell(row=(cur_row + 1), column=(i + 1)).value)
        return values


class MaxRowsReached(core.CoreExceptionError):
    """max rows reached exception"""

    def __init__(self, max_rows_count):
        super().__init__(f"maximum rows ({max_rows_count}) reached")
