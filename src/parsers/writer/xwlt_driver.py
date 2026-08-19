"""
write price list logic via xlsxwriter module
"""

from typing import Any

import openpyxl
from openpyxl.styles import Color, Font, PatternFill

from .ixls_driver import IXlsDriver

EXCEL_ALPHABET_SIZE = 26
EXCEL_COLUMN_A_ORD = 65


class WorkbookNotInitializedError(RuntimeError):
    """Raised when the xlsx driver is used before init_workbook."""

    def __init__(self) -> None:
        super().__init__("workbook is not initialized")


class WorksheetNotInitializedError(RuntimeError):
    """Raised when the xlsx driver is used before add_sheet."""

    def __init__(self) -> None:
        super().__init__("worksheet is not initialized")


def number_to_excel_column(number: int) -> str:
    """
    Конвертирует номер колонки в символьное обозначение Excel
    1 -> A, 2 -> B, ..., 26 -> Z, 27 -> AA, и т.д.
    """
    column_label = ""
    while number > 0:
        number, remainder = divmod(number - 1, EXCEL_ALPHABET_SIZE)
        column_label = chr(EXCEL_COLUMN_A_ORD + remainder) + column_label
    return column_label


def solid_fill(_color: str) -> PatternFill:
    """Solid PatternFill from #RRGGBB / RRGGBB."""
    rgb = Color(rgb=_color.lstrip("#"))
    return PatternFill(fgColor=rgb, fill_type="solid")


class XlsxWriterDriver(IXlsDriver):
    """
    write price list logic via openpyxl module
    """

    def __init__(self) -> None:
        """init"""
        self.work_book: openpyxl.Workbook | None = None
        self.work_sheet: Any | None = None
        self.current_col_index = 0
        self.current_row_index = 0
        self.col_max_length: dict[int, int] = {}
        self._file_name: str | None = None
        self.row_index_at = 1

    def init_workbook(self, _folder: str, _file_name: str) -> Any:
        if not self.work_book:
            self._file_name = _folder + _file_name
            self.work_book = openpyxl.Workbook()

        return self.work_book

    def add_sheet(self, sheet_name: str) -> XlsxWriterDriver:
        if self.work_book is None:
            raise WorkbookNotInitializedError()
        sheet = self.work_book.active
        sheet.title = sheet_name
        self.work_sheet = sheet
        return self

    def write_head(self, names: list[str]) -> None:
        """write head"""

        for col_idx, name in enumerate(names):
            self.write(0, col_idx, name, style=Font(bold=True))

    def set_column_format(self, column_format: dict[int, str]) -> None:
        """
        set column format
        :param column_format: dict[column_index, '@']
        """
        if self.work_sheet is None:
            raise WorksheetNotInitializedError()
        for index, c_format in column_format.items():
            self.work_sheet.column_dimensions[number_to_excel_column(index)].number_format = c_format

    def write(
        self,
        row_idx: int,
        col_idx: int,
        cell_content: Any,
        style: Any = None,
        _color: str | None = None,
    ) -> None:
        """write"""
        if self.work_sheet is None:
            raise WorksheetNotInitializedError()
        row_idx += self.row_index_at
        col_idx += self.row_index_at
        cell = self.work_sheet.cell(row=row_idx, column=col_idx, value=cell_content)
        if style:
            cell.font = style
        if _color:
            cell.fill = solid_fill(_color)

        self.current_col_index = col_idx
        self.current_row_index = row_idx
        content_length = len(str(cell_content or ""))
        known_max = self.col_max_length.get(col_idx)
        if known_max is None or known_max < content_length:
            self.col_max_length[col_idx] = content_length

    def save(self) -> None:
        """save file"""
        if self.work_sheet is None or self.work_book is None:
            raise WorkbookNotInitializedError()
        for col_index, max_len in self.col_max_length.items():
            self.work_sheet.column_dimensions[number_to_excel_column(col_index)].width = max_len + 4
        self.work_book.save(self._file_name)
        self.work_book.close()
